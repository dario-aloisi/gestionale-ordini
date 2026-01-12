import json
import os
import signal
import time
import threading
import logging
import shutil
import smtplib
import pandas as pd
import io # Serve per gestire il file in memoria RAM

from dotenv import load_dotenv
load_dotenv() # Carica le variabili dal file .env

from logging.handlers import RotatingFileHandler
from datetime import datetime, timedelta

# Importazioni per la creazione dell'excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Importazioni per la gestione delle Email (MIME)
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Importazioni Flask e Database
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, desc, extract
from sqlalchemy.exc import IntegrityError
from fpdf import FPDF

# Importazione Modelli dal file models.py
from models import db, Prodotto, Cliente, Ordine, DettaglioOrdine

app = Flask(__name__)

# ==============================================================================
# 1. CONFIGURAZIONE LOGGING (LA "SCATOLA NERA")
# ==============================================================================
# Questo sistema scrive gli errori gravi su un file 'errori.log' invece di perderli.
if not os.path.exists('logs'):
    try:
        os.makedirs('logs', exist_ok=True)
    except:
        pass # Se non riesce a creare la cartella logs, userà la root

# Configurazione del file di log (max 100KB, ne tiene 1 di backup)
handler = RotatingFileHandler('errori.log', maxBytes=100000, backupCount=1)
handler.setLevel(logging.ERROR) 
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
app.logger.addHandler(handler)

# ==============================================================================
# 2. CONFIGURAZIONI APP E DATABASE
# ==============================================================================
# Ora la chiave segreta la prende dal file .env
app.secret_key = os.getenv('SECRET_KEY', 'chiave_di_riserva_se_manca_env')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///gestionale.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

# ==============================================================================
# 3. CONFIGURAZIONE EMAIL (SICURO)
# ==============================================================================
# Legge dal file .env. Se non trova nulla, restituisce None
EMAIL_MITTENTE = os.getenv('EMAIL_MITTENTE')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD')
EMAIL_DESTINATARIO = os.getenv('EMAIL_DESTINATARIO')

# ==============================================================================
# 4. ROTTE PRINCIPALI
# ==============================================================================

@app.route('/')
def home():
    """
    Pagina iniziale (Dashboard).
    Calcola i numeri da mostrare nelle card colorate.
    """
    try:
        # Conta solo i clienti e prodotti attivi (non quelli cancellati)
        num_clienti = Cliente.query.filter_by(attivo=True).count()
        num_prodotti = Prodotto.query.filter_by(attivo=True).count()
        
        # Calcola gli ordini fatti nel mese corrente
        mese_corrente = datetime.now().month
        anno_corrente = datetime.now().year
        
        num_ordini_mese = Ordine.query.filter(
            extract('month', Ordine.data_consegna) == mese_corrente,
            extract('year', Ordine.data_consegna) == anno_corrente
        ).count()

        return render_template('home.html', n_cli=num_clienti, n_prod=num_prodotti, n_ord=num_ordini_mese)
    except Exception as e:
        app.logger.error(f"Errore caricamento HOME: {e}")
        return f"Errore caricamento Home: {e}", 500

# ==============================================================================
# 5. GESTIONE PRODOTTI
# ==============================================================================

@app.route('/prodotti')
def gestione_prodotti():
    # Mostra la lista dei prodotti (SOLO quelli attivi)
    lista_prodotti = Prodotto.query.filter_by(attivo=True).all()
    return render_template('prodotti.html', prodotti=lista_prodotti)

@app.route('/prodotti/aggiungi', methods=['POST'])
def aggiungi_prodotto():
    try:
        codice = request.form.get('codice')
        nome = request.form.get('nome')
        ingredienti = request.form.get('ingredienti')
        # 'next' ci dice se veniamo dalla pagina Ordine o dalla Gestione
        next_page = request.form.get('next')
        try:
            prezzo = float(request.form.get('prezzo', 0.0).replace(',', '.'))
        except:
            prezzo = 0.0

        if not codice or not nome:
            flash('Codice e Nome sono obbligatori!', 'error')
            return redirect(url_for('gestione_prodotti'))

        nuovo = Prodotto(codice=codice, nome=nome, ingredienti=ingredienti, prezzo=prezzo)
        db.session.add(nuovo)
        db.session.commit()
        flash('Prodotto aggiunto con successo!', 'success')
        
        # Se venivamo dall'ordine, torniamo lì
        if next_page == 'crea_ordine':
            return redirect(url_for('crea_ordine'))
        return redirect(url_for('gestione_prodotti', new_code=codice))

    except IntegrityError:
        # Gestisce il caso di Codice Duplicato
        db.session.rollback()
        flash(f"Errore: Il codice '{codice}' è già usato!", 'error')
        if request.form.get('next') == 'crea_ordine':
            # Riapre il modale se eravamo nell'ordine
            return redirect(url_for('crea_ordine', open_modal='prodotto'))
        return redirect(url_for('gestione_prodotti'))
    except Exception as e:
        app.logger.error(f"Errore AGGIUNGI PRODOTTO: {e}")
        flash(f"Errore di sistema: {e}", 'error')
        return redirect(url_for('gestione_prodotti'))

@app.route('/prodotti/modifica/<int:id_prodotto>', methods=['GET', 'POST'])
def modifica_prodotto(id_prodotto):
    try:
        prodotto = Prodotto.query.get_or_404(id_prodotto)
        # Gestione del parametro "next" (per sapere dove tornare)
        next_page = request.args.get('next')

        if request.method == 'POST':
            try:
                # Prendiamo i nuovi dati dal form
                nuovo_codice = request.form.get('codice')
                prodotto.codice = nuovo_codice
                prodotto.nome = request.form.get('nome')
                prodotto.ingredienti = request.form.get('ingredienti')
                try:
                    prodotto.prezzo = float(request.form.get('prezzo', 0.0).replace(',', '.'))
                except:
                    prodotto.prezzo = 0.0
                
                db.session.commit()
                flash('Prodotto modificato con successo!', 'success')

                # --- LOGICA DI RITORNO ---
                if next_page == 'crea_ordine':
                    # Se torno all'ordine, passo un parametro specifico per dire "seleziona questo prodotto"
                    return redirect(url_for('crea_ordine', updated_product=nuovo_codice))
                else:
                    # Se torno alla lista, passo new_code così la tabella scrolla e illumina
                    return redirect(url_for('gestione_prodotti', new_code=nuovo_codice))

            except IntegrityError:
                db.session.rollback()
                codice_errato = request.form.get('codice')
                flash(f"Errore: Il codice '{codice_errato}' è già usato!", 'error')
        
        return render_template('modifica_prodotto.html', prodotto=prodotto, next_page=next_page)

    except Exception as e:
        app.logger.error(f"Errore MODIFICA PRODOTTO id {id_prodotto}: {e}")
        return f"Errore interno: {e}", 500

@app.route('/prodotti/elimina/<int:id_prodotto>')
def elimina_prodotto(id_prodotto):
    """
    SOFT DELETE: Non cancella la riga, ma la nasconde (attivo=False).
    Rinomina il codice (aggiungendo timestamp) per liberarlo per nuovi usi.
    """
    try:
        prodotto = Prodotto.query.get_or_404(id_prodotto)
        prodotto.attivo = False
        
        # Liberiamo il codice (es. 100 -> 100_DEL_12345)
        suffisso = f"_DEL_{int(time.time())}"
        prodotto.codice = f"{prodotto.codice}{suffisso}"
        prodotto.nome = f"{prodotto.nome} (ELIMINATO)"
        
        db.session.commit()
        flash('Prodotto archiviato con successo.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Errore ELIMINA PRODOTTO id {id_prodotto}: {e}")
        flash('Errore durante l\'eliminazione.', 'error')
    
    return redirect(url_for('gestione_prodotti'))

# ==============================================================================
# 6. GESTIONE CLIENTI
# ==============================================================================

@app.route('/clienti')
def gestione_clienti():
    lista_clienti = Cliente.query.filter_by(attivo=True).all()
    return render_template('clienti.html', clienti=lista_clienti)

@app.route('/clienti/aggiungi', methods=['POST'])
def aggiungi_cliente():
    try:
        codice = request.form.get('codice')
        nome = request.form.get('nome')
        note = request.form.get('note')
        next_page = request.form.get('next')

        nuovo_cliente = Cliente(codice=codice, nome=nome, note=note)
        db.session.add(nuovo_cliente)
        db.session.commit()
        flash('Cliente aggiunto con successo!', 'success')

        if next_page == 'crea_ordine':
            return redirect(url_for('crea_ordine'))
        return redirect(url_for('gestione_clienti', new_code=codice))

    except IntegrityError:
        db.session.rollback()
        flash(f"Errore: Il codice '{codice}' esiste già!", 'error')
        if request.form.get('next') == 'crea_ordine':
            return redirect(url_for('crea_ordine', open_modal='cliente'))
        return redirect(url_for('gestione_clienti'))
    except Exception as e:
        app.logger.error(f"Errore AGGIUNGI CLIENTE: {e}")
        flash(f"Errore di sistema: {e}", 'error')
        return redirect(url_for('gestione_clienti'))

@app.route('/clienti/modifica/<int:id_cliente>', methods=['GET', 'POST'])
def modifica_cliente(id_cliente):
    try:
        cliente = Cliente.query.get_or_404(id_cliente)
        next_page = request.args.get('next')

        if request.method == 'POST':
            try:
                nuovo_codice = request.form.get('codice')
                cliente.codice = nuovo_codice
                cliente.nome = request.form.get('nome')
                cliente.note = request.form.get('note')
                
                db.session.commit()
                flash('Cliente modificato con successo!', 'success')

                # --- LOGICA DI RITORNO ---
                if next_page == 'crea_ordine':
                    return redirect(url_for('crea_ordine', updated_client=nuovo_codice))
                else:
                    return redirect(url_for('gestione_clienti', new_code=nuovo_codice))

            except IntegrityError:
                db.session.rollback()
                flash(f"Errore: Il codice '{request.form.get('codice')}' è già usato!", 'error')

        return render_template('modifica_cliente.html', cliente=cliente, next_page=next_page)

    except Exception as e:
        app.logger.error(f"Errore MODIFICA CLIENTE id {id_cliente}: {e}")
        return f"Errore interno: {e}", 500

@app.route('/clienti/elimina/<int:id_cliente>')
def elimina_cliente(id_cliente):
    """
    SOFT DELETE: Non cancella la riga, ma la nasconde (attivo=False).
    Rinomina il codice (aggiungendo timestamp) per liberarlo per nuovi usi.
    """
    try:
        cliente = Cliente.query.get_or_404(id_cliente)
        cliente.attivo = False
        # Liberiamo il codice (es. 100 -> 100_DEL_12345)
        suffisso = f"_DEL_{int(time.time())}"
        cliente.codice = f"{cliente.codice}{suffisso}"
        cliente.nome = f"{cliente.nome} (ELIMINATO)"
        
        db.session.commit()
        flash('Cliente archiviato con successo.', 'success')
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Errore ELIMINA CLIENTE id {id_cliente}: {e}")
        flash('Errore durante l\'eliminazione.', 'error')
    
    return redirect(url_for('gestione_clienti'))

# ==============================================================================
# 7. CREAZIONE ORDINE E PDF
# ==============================================================================

@app.route('/crea_ordine')
def crea_ordine():
    # Passiamo solo clienti e prodotti ATTIVI alle select
    tutti_clienti = Cliente.query.filter_by(attivo=True).all()
    tutti_prodotti = Prodotto.query.filter_by(attivo=True).all()
    open_modal = request.args.get('open_modal')
    return render_template('crea_ordine.html', clienti=tutti_clienti, prodotti=tutti_prodotti, open_modal=open_modal)

def pulisci_testo(testo):
    """Pulisce i caratteri strani per evitare crash del PDF"""
    if not testo: return ""
    return testo.encode('latin-1', 'replace').decode('latin-1')

class PDF(FPDF):
    def header(self):
        self.set_font('Helvetica', 'B', 14)
        self.cell(0, 10, 'Agente 15 ALOISI GIANCARLO', align='C', new_x="LMARGIN", new_y="NEXT")
        self.ln(1) # Spazio dopo l'intestazione

@app.route('/api/suggerimenti_cliente/<int:cliente_id>')
def api_suggerimenti_cliente(cliente_id):
    # Trova i prodotti comprati da questo cliente e quante volte
    # Restituisce: {id_prodotto: totale_acquistato}
    risultati = db.session.query(
        DettaglioOrdine.prodotto_id,
        func.sum(DettaglioOrdine.quantita).label('totale')
    ).join(Ordine).filter(
        DettaglioOrdine.cliente_id == cliente_id,
        Ordine.stato == 'inviato' # Consideriamo solo ordini confermati? O tutti? Meglio tutti per sicurezza
    ).group_by(DettaglioOrdine.prodotto_id).all()
    
    # Creiamo un dizionario semplice { "ID_PRODOTTO": TOTALE }
    suggerimenti = {str(r.prodotto_id): r.totale for r in risultati}
    return jsonify(suggerimenti)

@app.route('/genera_anteprima', methods=['POST'])
def genera_anteprima():
    """
    Riceve i dati dal frontend, crea il PDF nella cartella temp 
    e salva i dati in Sessione per l'invio successivo.
    """
    try:
        # 1. Creazione cartella temp e pulizia vecchi file
        percorso_temp = os.path.join(app.root_path, 'static', 'temp')
        if not os.path.exists(percorso_temp):
            os.makedirs(percorso_temp)
        
        for filename in os.listdir(percorso_temp):
            file_path = os.path.join(percorso_temp, filename)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                print(f"Errore pulizia: {e}")

        # 2. Riceviamo i dati e li SALVIAMO IN SESSIONE
        dati_json = request.get_json()
        session['dati_ordine_temp'] = dati_json

        raw_data = dati_json.get('data', 'N/D')
        note_generali = dati_json.get('note', '')
        righe = dati_json.get('righe', [])

        # Parsing Data (Doppio formato: per PDF e per Nome File)
        try:
            date_obj = datetime.strptime(raw_data, '%Y-%m-%d')
            data_per_pdf = date_obj.strftime('%d/%m/%Y')
            data_per_filename = date_obj.strftime('%d-%m-%Y')
        except ValueError:
            data_per_pdf = raw_data
            data_per_filename = "senza_data"

        # 3. Trasformazione Dati in Matrice per il PDF
        clienti_header = {} 
        prodotti_matrix = {} 
        totali_per_cliente = {} 

        for riga in righe:
            c_id = str(riga['cliente_id'])
            p_id = str(riga['prodotto_id'])
            qta = int(riga['quantita'])
            
            # Parsing Cliente
            c_raw = riga.get('cliente_check', 'Sconosciuto')
            if ' (' in c_raw:
                parts = c_raw.split(' (Cod. ')
                c_nome = parts[0]
                c_cod = parts[1].replace(')', '') if len(parts) > 1 else "N/D"
            else:
                c_nome = c_raw
                c_cod = "N/D"
            
            if c_id not in clienti_header:
                clienti_header[c_id] = {'nome': pulisci_testo(c_nome), 'codice': pulisci_testo(c_cod)}
                totali_per_cliente[c_id] = 0

            totali_per_cliente[c_id] += qta

            # Parsing Prodotto
            p_raw = riga.get('prodotto_check', 'Sconosciuto')
            if ' (' in p_raw:
                parts = p_raw.split(' (Cod. ')
                p_nome = parts[0]
                p_cod = parts[1].replace(')', '') if len(parts) > 1 else "N/D"
            else:
                p_nome = p_raw
                p_cod = "N/D"

            if p_id not in prodotti_matrix:
                prodotti_matrix[p_id] = {
                    'nome': pulisci_testo(p_nome),
                    'codice': pulisci_testo(p_cod),
                    'qta_clienti': {}
                }
            prodotti_matrix[p_id]['qta_clienti'][c_id] = qta

        # 4. Creazione PDF Grafico SU MISURA
        
        # --- A. DEFINIZIONE MISURE ---
        w_cod_prod = 20   
        w_nom_prod = 70   
        w_cli = 22        
        w_tot = 15        
        h_row = 7         
        margine_laterale = 10

        # --- B. CALCOLO LARGHEZZA ---
        # MargineSX + Cod + Nome + (NumeroClienti * LarghezzaCliente) + Totale + MargineDX
        num_clienti = len(clienti_header)
        width_custom = margine_laterale + w_cod_prod + w_nom_prod + (num_clienti * w_cli) + w_tot + margine_laterale

        # --- C. CALCOLO ALTEZZA ---
        
        # 1. Calcoliamo il numero di prodotti (righe)
        num_prodotti = len(prodotti_matrix)

        # 2. Altezza della struttura base (Intestazione + Tabella + Totali)
        # 30mm (Header Pagina) + 2 righe (Header Tabella) + N righe prodotti + 1 riga Totali
        height_base = 30 + (2 * h_row) + (num_prodotti * h_row) + h_row

        # 3. Calcolo dinamico spazio Note
        h_note = 0
        if note_generali and len(str(note_generali).strip()) > 0:
            # Titolo "Note Aggiuntive:" occupa circa 8mm
            # Stima: circa 90 caratteri stanno in una riga (con font size 9)
            lunghezza_testo = len(str(note_generali))
            num_righe_stimate = (lunghezza_testo // 90) + 1
            
            # 8mm per il titolo + (5mm * numero righe)
            h_note = 8 + (num_righe_stimate * 5)

        # 4. Altezza Totale + un margine di sicurezza in basso (es. 15mm)
        height_custom = height_base + h_note + 5

        # --- D. MINIMI DI SICUREZZA (Estetica) ---
        # Se è troppo piccolo, diamo una dimensione minima da foglio A5 orizzontale per non farlo sembrare uno scontrino
        #app.logger.error(f"Larghezza PDF calcolata: {width_custom}mm")
        if width_custom < 200: width_custom = 200
        #app.logger.error(f"Altezza PDF calcolata: {height_custom}mm")
        if height_custom < 80: height_custom = 80

        # --- E. ISTANZA PDF DINAMICA ---
        # Passiamo una tupla (larghezza, altezza) al posto di 'A4'
        pdf = PDF(orientation='P', unit='mm', format=(width_custom, height_custom))
        pdf.set_margins(10, 2, 10) #margini: sx, top, dx
        pdf.set_auto_page_break(auto=False) #margine inferiore entro cui creare un'altra pagina, se true
        pdf.add_page()
        pdf.set_font("Helvetica", 'B', size=10)

        # Testata
        pdf.cell(0, 6, f"Riepilogo del {data_per_pdf}", new_x="LMARGIN", new_y="NEXT")
        num_ordini = len(clienti_header)
        pdf.cell(0, 6, f"Numero ordini: {num_ordini}", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)     

        # --- TABELLA (Preparazione Dati) ---
        clienti_ordinati = sorted(clienti_header.items(), key=lambda x: x[1]['nome'])
        ids_clienti_ordinati = [x[0] for x in clienti_ordinati]

        # Calcolo larghezza totale tabella (serve per la cornice finale)
        total_table_width = w_cod_prod + w_nom_prod + (len(clienti_ordinati) * w_cli) + w_tot

        # MEMORIZZA Y INIZIO (Per la cornice)
        y_inizio_tabella = pdf.get_y()
        x_inizio_tabella = pdf.get_x()

        # Impostiamo linea SOTTILE per la griglia interna
        pdf.set_line_width(0.1) 
        pdf.set_font("Helvetica", 'B', 8)

        # RIGA 1 INTESTAZIONE
        pdf.cell(w_cod_prod, h_row, "", border=1)
        pdf.cell(w_nom_prod, h_row, "Cod. Cliente", border=1, align='C')
        for c_id, dati_c in clienti_ordinati:
            pdf.cell(w_cli, h_row, dati_c['codice'], border=1, align='C', fill=False)
        pdf.cell(w_tot, h_row, "TOT", border=1, align='C', new_x="LMARGIN", new_y="NEXT") 

        # RIGA 2 INTESTAZIONE
        pdf.cell(w_cod_prod, h_row, "Cod. Prod.", border=1, align='C')
        pdf.cell(w_nom_prod, h_row, "Nome Prodotto", border=1, align='C')
        for c_id, dati_c in clienti_ordinati:
            nome = dati_c['nome']
            display_nome = (nome[:10] + '.') if len(nome) > 10 else nome
            pdf.cell(w_cli, h_row, display_nome, border=1, align='C')
        pdf.cell(w_tot, h_row, "", border=1, align='C', new_x="LMARGIN", new_y="NEXT")

        # CORPO TABELLA
        pdf.set_font("Helvetica", 'B' ,size=8)
        totale_globale_cartoni = 0

        for p_id, dati in sorted(prodotti_matrix.items(), key=lambda x: x[1]['nome']):
            pdf.set_font("Helvetica", 'B' ,size=8)
            pdf.cell(w_cod_prod, h_row, dati['codice'], border=1, align='C')
            pdf.set_font("Helvetica",size=8)
            nome_p = (dati['nome'][:38] + '..') if len(dati['nome']) > 38 else dati['nome']
            pdf.cell(w_nom_prod, h_row, nome_p, border=1, align='L')
            pdf.set_font("Helvetica", 'B' ,size=8)
            
            totale_riga = 0
            for c_id in ids_clienti_ordinati:
                qta = dati['qta_clienti'].get(c_id, '') 
                display_qta = str(qta) if qta else '-'
                if qta: totale_riga += int(qta)
                
                if qta: pdf.set_font("Helvetica", 'B', 8)
                pdf.cell(w_cli, h_row, display_qta, border=1, align='C')
                if qta: pdf.set_font("Helvetica", 'B', size=8)
            
            pdf.set_font("Helvetica", 'B', 8)
            pdf.cell(w_tot, h_row, str(totale_riga), border=1, align='C', new_x="LMARGIN", new_y="NEXT")
            pdf.set_font("Helvetica", 'B', size=8)
            totale_globale_cartoni += totale_riga

        # RIGA TOTALI FINALI
        pdf.set_font("Helvetica", 'B', 8)
        pdf.cell(w_cod_prod, h_row, "", border=1)
        pdf.cell(w_nom_prod, h_row, "TOTALI", border=1, align='C')
        
        for c_id in ids_clienti_ordinati:
            somma_colonna = totali_per_cliente[c_id]
            pdf.cell(w_cli, h_row, str(somma_colonna), border=1, align='C')
            
        pdf.cell(w_tot, h_row, str(totale_globale_cartoni), border=1, align='C', new_x="LMARGIN", new_y="NEXT")

        # --- DISEGNO CORNICE ESTERNA SPESSA ---
        y_fine_tabella = pdf.get_y()
        pdf.set_line_width(0.4) 
        pdf.rect(x_inizio_tabella, y_inizio_tabella, total_table_width, y_fine_tabella - y_inizio_tabella)

        # NOTE PIE' DI PAGINA
        if note_generali and note_generali.strip():
            pdf.ln(2) # Un po' di spazio dalla tabella
            
            # 1. Scriviamo il titolo "Note Aggiuntive:" in grassetto
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(0, 5, "Note Aggiuntive:", new_x="LMARGIN", new_y="NEXT", align='L')
            
            # 2. Scriviamo il contenuto delle note in normale, andando a capo automaticamente
            pdf.set_font("Helvetica", "", 9)
            # 0 = larghezza totale pagina
            # 5 = altezza di ogni riga
            pdf.multi_cell(0, 5, pulisci_testo(note_generali), border=0, align='L')

        # Salvataggio su file temp

        # Aggiungiamo l'orario al nome file (Es. 10-30)
        orario = datetime.now().strftime('%H-%M')

        # Nome: preview_ordini_29-11-2025_orario_10-30.pdf
        nome_file = f"preview_ordini_{data_per_filename}_orario_{orario}.pdf"
        percorso_pdf = os.path.join(percorso_temp, nome_file)
        pdf.output(percorso_pdf)
        return jsonify({"status": "OK", "filename": nome_file})

    except Exception as e:
        print(f"ERRORE PDF: {e}")
        import traceback
        traceback.print_exc()
        app.logger.error(f"Errore GENERAZIONE PDF: {str(e)}")
        return str(e), 500

@app.route('/mostra_preview')
def mostra_preview():
    nome_file = request.args.get('file', 'preview_ordini.pdf')
    return render_template('preview.html', filename=nome_file)

# ==============================================================================
# 8. INVIO DEFINITIVO E SALVATAGGIO DB
# ==============================================================================

@app.route('/invia_definitivo', methods=['POST'])
def invia_definitivo():
    try:
        # 1. Recupero Dati
        dati_req = request.get_json()
        nome_file_preview = dati_req.get('filename')
        dati_ordine = session.get('dati_ordine_temp')

        # --- CONTROLLO DI SICUREZZA ---
        # Se la sessione è scaduta o vuota, ci fermiamo SUBITO.
        if not dati_ordine:
            return jsonify({
                "status": "KO", 
                "errore": "Sessione scaduta. I dati dell'ordine sono persi. Riprova a creare l'ordine."
            }), 400
        
        # --- ESTRAZIONE ORARIO DAL NOME FILE ---
        # Il nome è tipo: preview_ordini_DATA_ORARIO.pdf
        # Splittiamo per "_" e prendiamo l'ultimo pezzo
        parti_nome = nome_file_preview.split('_') 
        # L'ultimo pezzo è "10-30.pdf"
        orario_sporco = parti_nome[-1] 
        orario_pulito = orario_sporco.replace('.pdf', '') # "10-30"

        # 2. Salvataggio nel Database (Siamo sicuri che i dati sono in sessione)
        try:
            data_obj = datetime.strptime(dati_ordine.get('data'), '%Y-%m-%d').date()
            
            nuovo_ordine = Ordine(
                data_consegna=data_obj,
                note=dati_ordine.get('note'),
                stato='inviato',
                ora_creazione=orario_pulito
            )
            db.session.add(nuovo_ordine)
            db.session.flush() # Otteniamo l'ID
            
            for riga in dati_ordine.get('righe', []):
                # --- RECUPERO PREZZO ATTUALE ---
                prod_db = Prodotto.query.get(riga['prodotto_id'])
                prezzo_unit = prod_db.prezzo if prod_db else 0.0
                
                dettaglio = DettaglioOrdine(
                    ordine_id=nuovo_ordine.id,
                    cliente_id=int(riga['cliente_id']),
                    prodotto_id=int(riga['prodotto_id']),
                    quantita=int(riga['quantita']),
                    prezzo_storico=prezzo_unit # Prezzo al momento dell'ordine
                )
                db.session.add(dettaglio)
            
            db.session.commit()
            print("--- ORDINE SALVATO NEL DATABASE ---")
            
            # Puliamo la sessione SOLO DOPO aver salvato
            session.pop('dati_ordine_temp', None)
            
        except Exception as e_db:
            db.session.rollback()
            app.logger.error(f"ERRORE SALVATAGGIO DB: {e_db}")
            print(f"ERRORE SALVATAGGIO DB: {e_db}")
            # Se fallisce il DB, ci fermiamo? O inviamo lo stesso?
            # Meglio fermarsi per coerenza dati.
            return jsonify({"status": "KO", "errore": f"Errore Salvataggio DB: {str(e_db)}"}), 500
        
        # 3. Gestione File PDF (Spostamento in Archivio)
        cartella_preview = os.path.join(app.root_path, 'static', 'temp')
        path_preview = os.path.join(cartella_preview, nome_file_preview)
        
        # A. NOME PER L'ARCHIVIO LOCALE (Manteniamo l'orario per unicità)
        # Es: ordini_15-12-2025_orario_10-30.pdf
        nome_file_archivio = nome_file_preview.replace('preview_ordini', 'ordini')
        
        # B. NOME PER L'EMAIL (Pulito, solo data)
        # Es: ordini_15-12-2025.pdf
        # Ricaviamo la data pulita dall'oggetto ordine o dal nome file
        data_pulita = nome_file_preview.split('_')[2] # Prende la parte della data (es. 15-12-2025)
        nome_file_email = f"ordini_{data_pulita}.pdf"
        
        # 3. SALVATAGGIO LOCALE
        cartella_archivio = os.path.join(app.root_path, 'ARCHIVIO_PDF') 
        if not os.path.exists(cartella_archivio):
            os.makedirs(cartella_archivio)
            
        path_archivio = os.path.join(cartella_archivio, nome_file_archivio)
        shutil.copy(path_preview, path_archivio)
        print(f"File archiviato in: {path_archivio}")

        # 4. Invio Email
        msg = MIMEMultipart()
        msg['From'] = EMAIL_MITTENTE
        msg['To'] = EMAIL_DESTINATARIO
        msg['Subject'] = f"Consegne del {data_pulita}"

        body = ""
        msg.attach(MIMEText(body, 'plain'))

        # Allegato PDF
        with open(path_archivio, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
        
        encoders.encode_base64(part)

        # Alleghiamo il contenuto fisico di 'path_archivio' (che ha l'orario)
        # Ma diciamo alla mail di chiamarlo 'nome_file_email' (senza orario)
        part.add_header("Content-Disposition", f"attachment; filename= {nome_file_email}")
        msg.attach(part)

        # INVIO MAIL CON SMTP
        server = smtplib.SMTP_SSL('smtp.mail.yahoo.com', 465)
        server.login(EMAIL_MITTENTE, EMAIL_PASSWORD)
        server.send_message(msg)
        server.quit()

        return jsonify({
            "status": "OK", 
            "messaggio": "Email inviata e file archiviato!",
            "id_ordine": nuovo_ordine.id if 'nuovo_ordine' in locals() else 0
        })

    except Exception as e:
        app.logger.error(f"Errore INVIO EMAIL/ARCHIVIO: {str(e)}")
        print(f"ERRORE INVIO: {e}")
        return jsonify({"status": "KO", "errore": str(e)}), 500

# ==============================================================================
# 9. UTILITIES E STORICO
# ==============================================================================

def spegnimento_ritardato():
    time.sleep(1)
    print("--- SPEGNIMENTO ---")
    os.kill(os.getpid(), signal.SIGINT)

@app.route('/spegni', methods=['GET'])
def spegni_server():
    threading.Thread(target=spegnimento_ritardato).start()
    return "Bye"

@app.route('/backup_dati')
def backup_dati():
    try:
        # 1. Cerca il DB nella root o nella cartella instance
        db_path = os.path.join(app.root_path, 'gestionale.db')
        if not os.path.exists(db_path):
            db_path = os.path.join(app.root_path, 'instance', 'gestionale.db')

        if not os.path.exists(db_path):
            return jsonify({"status": "KO", "errore": "Database originale non trovato"}), 404
            
        # 2. Crea cartella BACKUP se non esiste
        backup_dir = os.path.join(app.root_path, 'BACKUP')
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir)

        # 3. Genera il nome del file (Data + Ora per evitare sovrascritture)
        nome_file = f"backup_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.db"
        dest_path = os.path.join(backup_dir, nome_file)

        # 4. Copia fisica del file
        import shutil
        shutil.copy2(db_path, dest_path)

        # 5. Risponde al Frontend con successo
        return jsonify({
            "status": "OK", 
            "messaggio": "Backup salvato correttamente!", 
            "path": dest_path,
            "filename": nome_file
        })

    except Exception as e:
        app.logger.error(f"Errore BACKUP: {e}")
        return jsonify({"status": "KO", "errore": str(e)}), 500

@app.route('/storico')
def storico():
    try:
        tutti_clienti = Cliente.query.filter_by(attivo=True).all()
        tutti_ordini = Ordine.query.order_by(Ordine.data_consegna.desc()).all()
        return render_template('storico.html', clienti=tutti_clienti, ordini=tutti_ordini)
    except Exception as e:
        app.logger.error(f"Errore caricamento STORICO: {e}")
        return f"Errore caricamento storico: {e}", 500

@app.route('/api/storico_cliente/<int:cliente_id>')
def api_storico_cliente(cliente_id):
    try:
        risultati = db.session.query(
            # Aggiungiamo Prodotto.id all'inizio per poter recuperare il prezzo specifico dopo
            Prodotto.id,
            Prodotto.codice,
            Prodotto.nome,
            func.sum(DettaglioOrdine.quantita).label('totale_pezzi'),
            func.max(Ordine.data_consegna).label('ultima_volta')
        ).join(Prodotto, DettaglioOrdine.prodotto_id == Prodotto.id)\
         .join(Ordine, DettaglioOrdine.ordine_id == Ordine.id)\
         .filter(DettaglioOrdine.cliente_id == cliente_id)\
         .group_by(Prodotto.id)\
         .order_by(desc('totale_pezzi'))\
         .all()

        data = []
        for r in risultati:
            # RECUPERO PREZZO: Cerchiamo il prezzo esatto di quel prodotto in quella data
            ultimo_prezzo = 0.0
            if r.ultima_volta:
                det = (DettaglioOrdine.query
                       .join(Ordine)
                       .filter(
                           DettaglioOrdine.cliente_id == cliente_id,
                           DettaglioOrdine.prodotto_id == r.id,
                           Ordine.data_consegna == r.ultima_volta
                       ).first())
                if det:
                    ultimo_prezzo = det.prezzo_storico

            data.append({
                'codice': r.codice,
                'nome': r.nome,
                'totale': r.totale_pezzi,
                'ultima_data': r.ultima_volta.strftime('%d/%m/%Y') if r.ultima_volta else '-',
                'ultimo_prezzo': ultimo_prezzo
            })
        return jsonify(data)

    except Exception as e:
        app.logger.error(f"Errore API STORICO CLIENTE: {e}")
        return jsonify([]), 500

@app.route('/api/dettaglio_ordine/<int:ordine_id>')
def api_dettaglio_ordine(ordine_id):
    try:
        ordine = Ordine.query.get_or_404(ordine_id)
        dettagli = DettaglioOrdine.query.filter_by(ordine_id=ordine_id).all()
        
        totale_ordine = 0.0
        totale_cartoni = 0        # <--- Contatore Cartoni
        clienti_unici = set()     # <--- Set per contare i clienti senza duplicati
        
        lista_righe = []
        for d in dettagli:
            prod = Prodotto.query.get(d.prodotto_id)
            cli = Cliente.query.get(d.cliente_id)
            
            # Usiamo il prezzo storico salvato nella riga (o 0.0 se mancante)
            prezzo_reale = d.prezzo_storico if d.prezzo_storico is not None else 0.0
            
            # Calcolo totale riga
            tot_riga = d.quantita * prezzo_reale
            totale_ordine += tot_riga
            
            # Aggiornamento statistiche
            totale_cartoni += d.quantita
            if cli:
                clienti_unici.add(cli.id)

            lista_righe.append({
                'cliente': cli.nome if cli else "Cliente Cancellato",
                'prodotto': prod.nome if prod else "Prodotto Cancellato",
                'codice': prod.codice if prod else "-",
                'quantita': d.quantita,
                'prezzo_unit': prezzo_reale, # <--- Ora manda il prezzo storico corretto
                'prezzo_tot': tot_riga 
            })
            
        response = {
            'info': {
                'data': ordine.data_consegna.strftime('%d/%m/%Y'),
                'note': ordine.note if ordine.note else "Nessuna nota",
                'totale_generale': totale_ordine,
                'num_cartoni': totale_cartoni,        # <--- Aggiunto
                'num_clienti': len(clienti_unici)     # <--- Aggiunto
            },
            'righe': lista_righe
        }
        return jsonify(response)

    except Exception as e:
        app.logger.error(f"Errore API DETTAGLIO ORDINE: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/modifica_ordine/<int:ordine_id>')
def modifica_ordine_page(ordine_id):
    ordine = Ordine.query.get_or_404(ordine_id)
    
    # 1. Carichiamo PRODOTTI e CLIENTI per le select di aggiunta/modifica
    prodotti = Prodotto.query.filter_by(attivo=True).order_by(Prodotto.nome).all()
    clienti_tutti = Cliente.query.filter_by(attivo=True).order_by(Cliente.nome).all()
    
    # 2. Prepariamo i dettagli attuali
    dettagli_list = []
    
    # Qui usiamo ordine.righe che è una lista di oggetti DettaglioOrdine
    for d in ordine.righe:
        prod = Prodotto.query.get(d.prodotto_id)
        cli = Cliente.query.get(d.cliente_id)
        
        dettagli_list.append({
            'cliente_id': d.cliente_id,
            'cliente_nome': cli.nome if cli else "Cliente Cancellato",
            'prod_id': d.prodotto_id,
            'prod_nome': prod.nome if prod else "Prodotto Cancellato",
            'qta': d.quantita,
            'prezzo': d.prezzo_storico if d.prezzo_storico is not None else 0.00
        })

    return render_template('modifica_ordine.html', 
                           ordine=ordine, 
                           prodotti=prodotti, 
                           clienti_tutti=clienti_tutti, # <--- Passiamo anche i clienti
                           dettagli_iniziali=dettagli_list)

@app.route('/api/salva_modifica_ordine', methods=['POST'])
def salva_modifica_ordine():
    try:
        data = request.json
        ordine_id = data.get('ordine_id')
        nuove_righe = data.get('righe') # Lista di {prod_id, qta, prezzo}

        ordine = Ordine.query.get_or_404(ordine_id)

        # 1. Cancelliamo TUTTI i vecchi dettagli di questo ordine
        # (È il metodo più sicuro per gestire rimozioni e modifiche insieme)
        DettaglioOrdine.query.filter_by(ordine_id=ordine_id).delete()

        # 2. Inseriamo le nuove righe
        # Nota: Dobbiamo recuperare il cliente_id per ogni riga.
        # In questo modello semplificato, assumiamo che l'ordine abbia un cliente principale?
        # ASPETTA: Nel tuo modello l'ordine è giornaliero e ogni riga ha il SUO cliente.
        # QUINDI: Nella modifica dobbiamo sapere per ogni riga chi è il cliente.
        # SE LA MODIFICA È "Togliere un prodotto a Mario", dobbiamo sapere che era di Mario.
        
        # FIX: Per semplificare la UI, assumo che tu stia modificando un ordine
        # ma NON stiamo cambiando i clienti delle righe esistenti, solo qta e prezzo.
        # PERÒ se aggiungi un prodotto nuovo, a chi lo assegniamo?
        # SOLUZIONE: Nel template ti farò vedere anche il Cliente di ogni riga.
        pass

        # RIVEDIAMO LA LOGICA DI SALVATAGGIO PER IL TUO MODELLO (Ordine Giornaliero Multi-Cliente)
        # Se modifichi una riga, devi rimandare al server anche il 'cliente_id'.
        
        for riga in nuove_righe:
            # riga = {'cliente_id': 1, 'prod_id': 5, 'qta': 10, 'prezzo': 5.50}
            nuovo_det = DettaglioOrdine(
                ordine_id=ordine_id,
                cliente_id=riga['cliente_id'],
                prodotto_id=riga['prod_id'],
                quantita=riga['qta'],
                prezzo_storico=riga['prezzo']
            )
            db.session.add(nuovo_det)

        # Aggiorniamo le note se modificate
        if 'note' in data:
            ordine.note = data['note']

        db.session.commit()
        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Errore Modifica Ordine: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/elimina_ordine/<int:ordine_id>', methods=['POST'])
def elimina_ordine(ordine_id):
    try:
        ordine = Ordine.query.get_or_404(ordine_id)
        
        # 1. Cancella prima tutti i dettagli (i prodotti dentro l'ordine)
        DettaglioOrdine.query.filter_by(ordine_id=ordine.id).delete()
        
        # 2. Cancella l'ordine principale
        db.session.delete(ordine)
        
        db.session.commit()
        return jsonify({'success': True})
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Errore eliminazione ordine {ordine_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/statistiche')
def api_statistiche():
    try:
        top_prodotti = db.session.query(
            Prodotto.nome,
            func.sum(DettaglioOrdine.quantita).label('totale')
        ).join(DettaglioOrdine).group_by(Prodotto.id).order_by(desc('totale')).limit(5).all()

        top_clienti = db.session.query(
            Cliente.nome,
            func.sum(DettaglioOrdine.quantita).label('totale')
        ).join(DettaglioOrdine).group_by(Cliente.id).order_by(desc('totale')).limit(5).all()

        vendite_mensili = db.session.query(
            func.strftime('%Y-%m', Ordine.data_consegna).label('mese'),
            func.sum(DettaglioOrdine.quantita).label('totale')
        ).join(DettaglioOrdine).group_by('mese').order_by('mese').all()

        oggi = datetime.now().date()
        soglia_dormienti = oggi - timedelta(days=30)

        query_dormienti = db.session.query(
            Cliente.nome,
            Cliente.id,
            func.max(Ordine.data_consegna).label('ultima_data')
        ).outerjoin(DettaglioOrdine, Cliente.id == DettaglioOrdine.cliente_id)\
         .outerjoin(Ordine, DettaglioOrdine.ordine_id == Ordine.id)\
         .filter(Cliente.attivo == True)\
         .group_by(Cliente.id).all()

        lista_dormienti = []
        for c_nome, c_id, ultima_data in query_dormienti:
            if ultima_data is None:
                lista_dormienti.append({'nome': c_nome, 'giorni': "Mai ordinato"})
            elif ultima_data < soglia_dormienti:
                delta = (oggi - ultima_data).days
                lista_dormienti.append({'nome': c_nome, 'giorni': f"{delta} giorni fa"})

        lista_dormienti.sort(key=lambda x: 9999 if x['giorni'] == "Mai ordinato" else int(x['giorni'].split()[0]), reverse=False)

        data = {
            'prodotti': {'labels': [r.nome for r in top_prodotti], 'values': [r.totale for r in top_prodotti]},
            'clienti': {'labels': [r.nome for r in top_clienti], 'values': [r.totale for r in top_clienti]},
            'andamento': {'labels': [r.mese for r in vendite_mensili], 'values': [r.totale for r in vendite_mensili]},
            'dormienti': lista_dormienti
        }
        return jsonify(data)
    except Exception as e:
        app.logger.error(f"Errore API STATISTICHE: {e}")
        return jsonify({}), 500

@app.route('/api/statistiche_economiche')
def statistiche_economiche():
    try:
        # 1. FATTURATO STORICO COMPLETO MENSILE (Tutti gli anni)
        # Raggruppiamo per "Anno-Mese" (es. "2025-11", "2025-12", "2026-01")
        fatturato_storico = db.session.query(
            func.strftime('%Y-%m', Ordine.data_consegna).label('anno_mese'),
            func.sum(DettaglioOrdine.quantita * DettaglioOrdine.prezzo_storico).label('totale')
        ).join(Ordine)\
         .filter(Ordine.stato != 'cancellato')\
         .group_by('anno_mese')\
         .order_by('anno_mese')\
         .all()

        # Prepariamo due liste dinamiche: Labels (Assi X) e Valori (Assi Y)
        mesi_labels = []
        mesi_values = []

        for r in fatturato_storico:
            # r.anno_mese arriva come "2025-11"
            # Lo trasformiamo in "11/2025" per renderlo leggibile
            parti = r.anno_mese.split('-') 
            label_leggibile = f"{parti[1]}/{parti[0]}"
            
            mesi_labels.append(label_leggibile)
            mesi_values.append(r.totale)

        # 2. TOP 10 CLIENTI PER FATTURATO
        top_clienti = db.session.query(
            Cliente.nome,
            func.sum(DettaglioOrdine.quantita * DettaglioOrdine.prezzo_storico).label('totale')
        ).join(DettaglioOrdine, DettaglioOrdine.cliente_id == Cliente.id)\
         .join(Ordine, DettaglioOrdine.ordine_id == Ordine.id)\
         .filter(Ordine.stato != 'cancellato')\
         .group_by(Cliente.id)\
         .order_by(desc('totale'))\
         .limit(10)\
         .all()

        # 3. TOP 10 PRODOTTI PER FATTURATO
        top_prodotti = db.session.query(
            Prodotto.nome,
            func.sum(DettaglioOrdine.quantita * DettaglioOrdine.prezzo_storico).label('totale')
        ).join(DettaglioOrdine, DettaglioOrdine.prodotto_id == Prodotto.id)\
         .join(Ordine, DettaglioOrdine.ordine_id == Ordine.id)\
         .filter(Ordine.stato != 'cancellato')\
         .group_by(Prodotto.id)\
         .order_by(desc('totale'))\
         .limit(10)\
         .all()

        return jsonify({
            'mesi_labels': mesi_labels,   # <--- Nuova lista dinamica
            'mesi_values': mesi_values,   # <--- Nuova lista valori
            'top_clienti': {
                'labels': [r.nome for r in top_clienti], 
                'values': [r.totale for r in top_clienti]
            },
            'top_prodotti': {
                'labels': [r.nome for r in top_prodotti], 
                'values': [r.totale for r in top_prodotti]
            }
        })

    except Exception as e:
        app.logger.error(f"Errore Statistiche Economiche: {e}")
        return jsonify({'error': str(e)}), 500

# ==============================================================================
# 10. SCARICA FOGLIO EXCEL DA DETTAGLI ORDINE PASSATI
# ==============================================================================

@app.route('/scarica_ordine_excel/<int:ordine_id>')
def scarica_ordine_excel(ordine_id):
    try:
        # 1. Recuperiamo i dati
        ordine = Ordine.query.get_or_404(ordine_id)
        dettagli = DettaglioOrdine.query.filter_by(ordine_id=ordine_id).all()

        # 2. Trasformazione Dati (Identica al PDF)
        clienti_header = {} 
        prodotti_matrix = {} 
        totali_per_cliente = {} 

        for d in dettagli:
            c_id = str(d.cliente_id)
            p_id = str(d.prodotto_id)
            qta = d.quantita
            
            # Info Cliente
            cli = Cliente.query.get(d.cliente_id)
            c_nome = cli.nome if cli else "Cancellato"
            c_cod = cli.codice if cli else "N/D"

            if c_id not in clienti_header:
                clienti_header[c_id] = {'nome': c_nome, 'codice': c_cod}
                totali_per_cliente[c_id] = 0
            
            totali_per_cliente[c_id] += qta

            # Info Prodotto
            prod = Prodotto.query.get(d.prodotto_id)
            p_nome = prod.nome if prod else "Cancellato"
            p_cod = prod.codice if prod else "N/D"

            if p_id not in prodotti_matrix:
                prodotti_matrix[p_id] = {
                    'nome': p_nome,
                    'codice': p_cod,
                    'qta_clienti': {}
                }
            prodotti_matrix[p_id]['qta_clienti'][c_id] = qta

        # Ordinamento Clienti e Prodotti
        clienti_ordinati = sorted(clienti_header.items(), key=lambda x: x[1]['nome'])
        ids_clienti_ordinati = [x[0] for x in clienti_ordinati]
        
        prodotti_ordinati = sorted(prodotti_matrix.items(), key=lambda x: x[1]['nome'])

        # 3. CREAZIONE EXCEL (OpenPyXL)
        wb = Workbook()
        ws = wb.active
        ws.title = "Riepilogo Ordine"

        # --- STILI ---
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

        # --- INTESTAZIONE DOCUMENTO ---
        ws['A1'] = "Agente 15 ALOISI GIANCARLO"
        ws['A1'].font = Font(bold=True, size=14)
        
        data_str = ordine.data_consegna.strftime('%d/%m/%Y')
        ws['A3'] = f"Riepilogo del {data_str}"
        ws['A3'].font = Font(bold=True)
        ws['A4'] = f"Numero ordini: {len(clienti_header)}"
        ws['A4'].font = Font(bold=True)

        row_idx = 6 # Iniziamo a disegnare la tabella dalla riga 6

        # --- INTESTAZIONE TABELLA (RIGA 1: Codici Cliente) ---
        # Col A: Cod. Prod, Col B: Nome Prod
        ws.cell(row=row_idx, column=2).value = "Cod. Cliente"
        ws.cell(row=row_idx, column=2).font = bold_font
        ws.cell(row=row_idx, column=2).alignment = center_align
        ws.cell(row=row_idx, column=2).border = thin_border

        col_idx = 3 # I clienti partono dalla colonna C (3)
        for c_id, dati_c in clienti_ordinati:
            c = ws.cell(row=row_idx, column=col_idx, value=dati_c['codice'])
            c.font = bold_font
            c.alignment = center_align
            c.border = thin_border
            col_idx += 1
        
        # Colonna TOTALE (in alto a destra)
        c_tot_h = ws.cell(row=row_idx, column=col_idx, value="TOT")
        c_tot_h.font = bold_font
        c_tot_h.alignment = center_align
        c_tot_h.border = thin_border
        
        row_idx += 1

        # --- INTESTAZIONE TABELLA (RIGA 2: Nomi) ---
        ws.cell(row=row_idx, column=1, value="Cod. Prod.").font = bold_font
        ws.cell(row=row_idx, column=1).border = thin_border
        ws.cell(row=row_idx, column=1).alignment = center_align

        ws.cell(row=row_idx, column=2, value="Nome Prodotto").font = bold_font
        ws.cell(row=row_idx, column=2).border = thin_border
        ws.cell(row=row_idx, column=2).alignment = center_align
        
        col_idx = 3
        for c_id, dati_c in clienti_ordinati:
            c = ws.cell(row=row_idx, column=col_idx, value=dati_c['nome'])
            c.alignment = center_align
            c.border = thin_border
            c.font = bold_font
            col_idx += 1
        
        # Cella vuota sotto TOT
        ws.cell(row=row_idx, column=col_idx).border = thin_border
        
        row_idx += 1

        # --- CORPO TABELLA ---
        totale_globale = 0

        for p_id, dati in prodotti_ordinati:
            # Codice Prodotto
            c1 = ws.cell(row=row_idx, column=1, value=dati['codice'])
            c1.font = bold_font
            c1.alignment = center_align
            c1.border = thin_border

            # Nome Prodotto
            c2 = ws.cell(row=row_idx, column=2, value=dati['nome'])
            #c2.font = bold_font
            c2.border = thin_border

            totale_riga = 0
            col_idx = 3
            
            for c_id in ids_clienti_ordinati:
                qta = dati['qta_clienti'].get(c_id, 0)
                valore_cella = qta if qta > 0 else "-"
                
                c = ws.cell(row=row_idx, column=col_idx, value=valore_cella)
                c.alignment = center_align
                c.border = thin_border
                if qta > 0:
                    c.font = bold_font
                    totale_riga += qta
                
                col_idx += 1
            
            # Totale Riga
            c_tot = ws.cell(row=row_idx, column=col_idx, value=totale_riga)
            c_tot.font = bold_font
            c_tot.alignment = center_align
            c_tot.border = thin_border
            
            totale_globale += totale_riga
            row_idx += 1

        # --- RIGA TOTALI FINALI ---
        ws.cell(row=row_idx, column=1).border = thin_border # Vuoto sotto codice
        
        c_label_tot = ws.cell(row=row_idx, column=2, value="TOTALI")
        c_label_tot.font = bold_font
        c_label_tot.alignment = center_align
        c_label_tot.border = thin_border

        col_idx = 3
        for c_id in ids_clienti_ordinati:
            somma = totali_per_cliente.get(c_id, 0)
            c = ws.cell(row=row_idx, column=col_idx, value=somma)
            c.font = bold_font
            c.alignment = center_align
            c.border = thin_border
            col_idx += 1
        
        c_grand_tot = ws.cell(row=row_idx, column=col_idx, value=totale_globale)
        c_grand_tot.font = bold_font
        c_grand_tot.alignment = center_align
        c_grand_tot.border = thin_border

        # --- NOTE ---
        if ordine.note:
            row_idx += 2
            ws.cell(row=row_idx, column=1, value=f"Note Aggiuntive: {ordine.note}")
            #ws.cell(row=row_idx, column=1).font = Font(italic=True)

        # --- FORMATTAZIONE LARGHEZZE COLONNE ---
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 40
        for i in range(3, col_idx + 1):
            ws.column_dimensions[get_column_letter(i)].width = 12

        # 4. Salvataggio in Memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 5. SALVATAGGIO SU DISCO
        cartella_excel = os.path.join(app.root_path, 'ARCHIVIO_EXCEL')
        if not os.path.exists(cartella_excel):
            os.makedirs(cartella_excel)

        data_str = ordine.data_consegna.strftime('%d-%m-%Y')
        # Generiamo il nome file usando l'ORARIO ORIGINALE DAL DB
        # Se per caso è un ordine vecchio senza orario, usiamo "00-00" o l'ora attuale come fallback
        orario_str = ordine.ora_creazione if ordine.ora_creazione else datetime.now().strftime('%H-%M')

        # Nome file identico al PDF locale
        nome_file = f"ordini_{data_str}_orario_{orario_str}.xlsx"
        
        path_completo = os.path.join(cartella_excel, nome_file)
        wb.save(path_completo) # Salva fisicamente nel progetto

        # Rispondiamo con un JSON per dire "Tutto ok"
        return jsonify({"status": "OK", "path": path_completo, "filename": nome_file})

    except Exception as e:
        app.logger.error(f"Errore creazione Excel: {e}")
        return jsonify({"status": "KO", "errore": str(e)}), 500

# ==============================================================================
# 11. AVVIO E FINE FILE
# ==============================================================================
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000)